#!/usr/bin/env python3
"""
Script one-shot : applique des fusions manuelles curatées sur merged-import.xlsx.
Chaque groupe liste les titres à fusionner. Le premier titre est le titre canonique.
Les données non-nulles des doublons sont mergées dans l'entrée canonique.
"""

import re
import unicodedata

import openpyxl

MERGED_PATH = "var/merged-import.xlsx"

def normalize(s):
    if s is None: return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[-'\u2019.,!?():]+", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# === MERGE MAP ===
# Format: { "sheet": [ ["canonical title", "dupe1", "dupe2", ...], ... ] }
# Le premier titre de chaque liste est conservé. Les suivants sont supprimés
# après fusion de leurs données.

MERGE_MAP = {
    "BD": [
        ["L'Agent 212", "Agent 212"],
        ["Les Armées du conquérant", "Les armées du conquerant", "Les Armées du conquérent"],
        ["Blake et Mortimer", "Blake & mortimer"],
        ["Les Chroniques de la lune noire", "Chroniques de la lune noire"],
        ["La Chute du dragon noir", "La Chute du dragon noir V1"],
        ["La Complainte des landes perdues", "Complainte des landes perdues"],
        ["Comptines d'Halloween", "Comptine d'Halloween"],
        ["C.O.P.S.", "C O P S", "C.O.P.S"],
        ["Le Donjon de Naheulbeuk", "Donjon de naheulbeuk", "Le Donjon de naheulbeuk"],
        ["L'Empreinte de Satan", "Empreinte de Satan", "L'Empreinte de satan"],
        ["Guerres parallèles", "Guerres paralleles", "Guerres paralleles - Peru"],
        ["La Horde du contrevent", "La Horde du contre vent"],
        ["I.R.$", "I R.$"],
        ["Johan et Pirlouit", "Johan & Pirlouit"],
        ["Les Lamentations de l'agneau", "Lamentations de l'agneau (les)"],
        ["La Licorne", "Licorne"],
        ["La Ligue des gentlemen extraordinaires", "La Ligue des gentleman extarordinaires",
         "la Ligue des gentlemen extraordinaires", "La Ligue des gentlemen extraordinaires - Century"],
        ["Le 3ème testament", "Le 3eme testament", "Le 3eme testament - Julius",
         "Le Troisieme testament", "Le Troisieme testament (integrale)", "Le Troisieme testament - julius"],
        ["Les Mondes d'Aldébaran", "Les Mondes d'aldebaran",
         "Les Mondes d'Aldebaran - Cycle 1 & 2 - Aldebaran & Betelgeuse"],
        ["Mjöllnir", "Mjollnir", "Mjollnir (legendes nordiques)"],
        ["Les Soleils rouges de l'éden", "Les Soleils rouge de l'eden", "Les Soleils rouges de l'eden"],
        ["Tueur de ville", "Tueur de ville - Vixit"],
        ["Uchronie(s) - New Byzance", "Uchronie new byzance"],
        ["Uchronie(s) - New York", "Uchronie new york"],
        ["Les Voleurs d'empires", "Les Voleurs d'empire"],
        ["Universal War One", "Universal war one", "Universal war two"],
        ["Crusades", "Crusaders"],
    ],
    "Comics": [
        ["100% Marvel - Le Projet Marvels", "100% Marvel - le projet marvel"],
        ["Alien vs Predator - Chasse à l'homme", "Alien vs predator - chasse à l'homme",
         "Aliens versus Predator - Chasse à l'homme [ Dark Horse ]"],
        ["Avengers - Celestial Quest", "Avengers - celestial quest",
         "Avengers - Celestial Quest (T01-08) (2001-2002)"],
        ["Bad Planet", "Bad planet", "Bad Planet (01-08) (2005-2013)"],
        ["Badrock and Company", "Badrock and company", "Badrock and Company (01-06)"],
        ["Badrock Annual", "Badrock annual", "Badrock Annual (Image)"],
        ["Badrock / Wolverine Savaged", "Badrock - wolverine savaged", "Badrock - Wolverine-Savaged"],
        ["Batman - Aliens 2", "Batman - aliens 2", "Batman - Aliens 2 (001-003)"],
        ["Captain Atom - Armageddon", "captain atom - armageddon 01 (of 09)", "captain atom - armageddon 02"],
        ["Cataclysm", "Cataclysm (story arc)", "Cataclysm - Story Arc (2013-2014)"],
        ["Civil War + Prologue", "Civil War & prologue", "Civil war + prologue"],
        ["Deadpool Killustrated", "Deadpool illustrated", "Deadpool Killustrated"],
        ["Divinity", "Divinity (T01-04)", "Divinity II", "Divinity III"],
        ["Flashback", "Flashback (Collection) (1997-2005)"],
        ["Hulk", "Hulk v3"],
        ["Infinity", "Infinity (FCBD)"],
        ["Infinity Watch", "Infinity watch", "Infinity Watch v01"],
        ["Injustice - Gods Among Us", "Injustice Gods Among Us",
         "Injustice gods among us - year one", "Injustice : Les dieux sont parmi nous"],
        ["Insexts", "Insexts (T1-8) (2015-2016)"],
        ["Justice League / Power Rangers", "Justice league - power rangers",
         "Justice League-Power Rangers 001"],
        ["Last Planet Standing", "Last planet standing", "Last Planet Standing (T01-05) 2006"],
        ["Leaving Megalopolis", "Leaving megalopolis", "Leaving Megalopolis Omnibus"],
        ["Lobo", "Lobo (001-013+Annual 01) (2014-) (digital) (Minutemen-Midas)",
         "Lobo - v1 (01-64+Annual+1000000) (1993-1999)"],
        ["Mage - The Hero Discovered", "Mage v1 - the hero discovered",
         "Mage v1 - The Hero Discovered (T01-15) (1984-1986)"],
        ["Magnus Robot Fighter", "Magnus robot fighter", "Magnus Robot Fighter Omnibus"],
        ["Mars Attacks", "Mars attacks", "Mars Attacks (Collection)", "Mars Attacks Image"],
        ["Marvel Universe - The End", "Marvel universe - the end",
         "Marvel Universe - The End (T01-06)"],
        ["Marvel Universe vs The Punisher", "Marvel universe vs the punisher",
         "Marvel Universe vs. The Punisher (T01-04)"],
        ["Marvel Zombies", "Marvel zombies", "Marvel Zombies  Saga"],
        ["Maximum Security", "Maximum security", "Maximum Security (T01-03) (2000-2001)"],
        ["Monsters, Myths and Marvels", "Monsters mythes and marvels",
         "Monsters, Myths and Marvels (T01-03)"],
        ["Moon Knight", "Moon Knight v1 30"],
        ["Nova Classic", "Nova classic", "Nova Classic Vol. 01"],
        ["Phoenix", "Phoenix (v1-v2) (1975-2012)"],
        ["Punisher", "Punisher (1986)", "Punisher (2011-2012)",
         "Punisher (T01-16) (2011-2012)", "Punisher v1 (T1-5)"],
        ["Punisher - In the Blood", "Punisher - in the blood",
         "Punisher - In the Blood (T01-05)"],
        ["Punisher MAX", "Punisher - max", "Punisher MAX (T01-22 + Annual + Special) (2008-2012)"],
        ["Punisher - Nightmare", "Punisher - nightmare", "Punisher Nightmare (T01-05)"],
        ["Punisher War Zone", "Punisher warzone",
         "Punisher - War Zone Vol 1 (1-41+Annual)", "Punisher - War Zone Vol. 2 (T1-6)"],
        ["Récit Complet Marvel", "Recit Complet", "Marvel récit complet"],
        ["Savage Dragon", "savage dragon 120"],
        ["Screwed", "Screwed (T1-6)"],
        ["Serval / Wolverine", "Serval - wolverine", "Serval - wolverine (1989 - 2010)"],
        ["Siege", "Siege (apres Secret Invasion)", "Siege (Collection)"],
        ["Silver Surfer", "Silver Surfer (v1 - v6 + Extras) (1968 - 2015)"],
        ["Silver Surfer - In Thy Name", "Silver surfer - in thy name",
         "Silver Surfer - In Thy Name (T01-04)"],
        ["StarCraft", "Starcraft", "StarCraft  (Comics Collection)"],
        ["Thanos", "Thanos (2003-2004)", "Thanos 001-012 (2003-2004)", "Thanos 2003 2004"],
        ["Thanos - Infinity Revelation/Relativity/Finale",
         "Thanos - infinity revelation relativity + final",
         "Thanos - Infinity Revelation, Relativity + Finale (2014-2016)"],
        ["Thanos Rising", "Thaos rising", "Thanos Rising (001-005)"],
        ["The Darkness", "The Darkness - Superman"],
        ["The Infinity Crusade", "The Infinity crusade", "The Infinity Crusade Omnibus"],
        ["The Mighty Thor", "The mighty thor", "The mighty Thor 2015",
         "The mighty thor 700-706", "The mighty Thor 700-706 2017"],
        ["The Punisher", "The Punisher (Marvel Max) (T01-T18)",
         "The Punisher v2", "The Punisher - Born (T01-04)"],
        ["The Royals - Masters of War", "The royales - masters of war",
         "The Royals - Masters of War 001-006"],
        ["The Witcher", "The Witcher intégrale"],
        ["Ultimate Secret", "Ultimate secret",
         "Ultimate Secret ( T 01 à 04 ) [ Intégrale ]"],
        ["Ultimates", "Ultimates 2", "Ultimates 3", "Ultimates 3 (001-005)", "Ultimates HS"],
        ["Universe X", "Universe x", "Universe X (00-05)"],
        ["Venom Space Knight", "Venom space knight 2015-2016",
         "Venom Space Knight (2015-2016) ANAD (C)"],
        ["Warlock and the Infinity Watch", "Warlock and the infinity watch",
         "Warlock and the Infinity Watch (T1-42) (1992-1995)"],
        ["Wetworks", "Wetworks (v1-v2+Extras) (1994-2010)"],
        ["WildC.A.T.s", "WildCATS", "WildC A T S"],
        ["Witchblade", "Witchblade (Delcourt)", "Witchblade (éditions USA)"],
        ["Wolverine (intégrale)", "Wolverine (integrale)", "Wolverine (integrale 1991)"],
        ["X-Men", "X-Men (Semic VI)", "X-Men V4", "X-men v4 (2013-2015)"],
        ["X-Men - Die by the Sword", "X-men  Die by Sword", "X-men die by the sword"],
        ["X-O Manowar", "X-O Manowar 01-03"],
        # Additional from user's list
        ["Heavy Metal (Marvel & Valiant Crossover)",
         "Heavy Metal (Marvel & Valiant Crossover) - Part 1 - X-O Manowar & Iron Man",
         "Heavy Metal (Marvel & Valiant) - Part 2 - Iron Man & X-O Manowar"],
        ["Promethée", "Promethee", "Promethee (v01-v06)"],
        ["Batman - Dark Knight", "Batman Dark Knight - Edition Integrale", "Batman - Dark Knight I"],
        ["Marvel Ultimate", "Marvel ultimate", "Marvel Ultimate vague 1"],
        ["Iron Man - Renaissance", "Iron man renaissance", "Iron Man Renaissance des héros"],
    ],
}


def main():
    wb = openpyxl.load_workbook(MERGED_PATH)
    total_merged = 0

    for sheet_name, groups in MERGE_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Build index: normalized title → row_idx
        title_index = {}  # norm_key → (row_idx, original_title)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            title = str(row[0].value).strip() if row[0].value else ""
            if title:
                key = normalize(title)
                title_index[key] = (row_idx, title)

        rows_to_delete = []

        for group in groups:
            canonical = group[0]
            dupes = group[1:]

            # Find all entries in this group
            found_entries = []
            for title in group:
                key = normalize(title)
                if key in title_index:
                    found_entries.append((title, key, title_index[key][0]))

            if len(found_entries) <= 1:
                continue  # Only 0-1 found, nothing to merge

            # Find canonical entry (or first found if canonical not in file)
            canonical_key = normalize(canonical)
            canonical_row = None
            for title, key, row_idx in found_entries:
                if key == canonical_key:
                    canonical_row = row_idx
                    break
            if canonical_row is None:
                # Canonical not found, use first entry as base
                canonical_row = found_entries[0][2]
                canonical_key = found_entries[0][1]

            # Set canonical title
            ws.cell(row=canonical_row, column=1).value = canonical

            # Merge data from dupes into canonical
            for title, key, row_idx in found_entries:
                if row_idx == canonical_row:
                    continue

                # Merge non-null values
                for col in range(2, 9):
                    if ws.cell(row=canonical_row, column=col).value is None:
                        dupe_val = ws.cell(row=row_idx, column=col).value
                        if dupe_val is not None:
                            ws.cell(row=canonical_row, column=col).value = dupe_val

                rows_to_delete.append(row_idx)
                total_merged += 1
                print(f"  [{sheet_name}] '{title}' → '{canonical}'")

        # Delete dupe rows (bottom to top)
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

    wb.save(MERGED_PATH)
    print(f"\n✓ {total_merged} doublons fusionnés dans {MERGED_PATH}")


if __name__ == "__main__":
    main()
