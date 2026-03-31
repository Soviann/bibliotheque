#!/usr/bin/env python3
"""
Fusionne Bibliotheque.xlsx + nas-import.xlsx + Livres.xlsx → import.xlsx
Enrichit les tomes achetés depuis Livres.xlsx dans le fichier fusionné.

Format de sortie : feuille unique "Import" avec colonne Type en premier.
"""

import difflib
import os
import re
import sys
import unicodedata

import openpyxl

# --- Paths ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BIBLIO_PATH = os.path.join(BASE_DIR, "var", "Bibliotheque.xlsx")
NAS_PATH = os.path.join(BASE_DIR, "var", "nas-import.xlsx")
LIVRES_PATH = os.path.join(BASE_DIR, "var", "Livres.xlsx")
OUTPUT_MERGED = os.path.join(BASE_DIR, "var", "import.xlsx")

# --- Column mappings ---
# Output format: Type(0), Titre(1), Buy?(2), Last bought(3), Current(4),
#                Parution(5), Last dled(6), On NAS?(7), Parution terminée(8),
#                ISBN(9), Auteur(10), Couverture(11), Éditeur(12), Catégories(13), Description(14)
BIBLIO_HEADERS = [
    "Type", "Titre", "Buy?", "Last bought", "Current",
    "Parution", "Last dled", "On NAS ?", "Parution terminée",
    "ISBN", "Auteur", "Couverture", "Éditeur", "Catégories", "Description",
]

# nas-import.xlsx columns → index in output (shifted +1 for Type column)
NAS_COL_MAP = {
    0: 1,  # Titre → Titre
    1: 2,  # Statut → Buy?
    2: 3,  # Dernier acheté → Last bought
    3: 4,  # Lu jusqu'à → Current
    4: 5,  # Nombre publié → Parution
    5: 6,  # Dernier téléchargé → Last dled
    6: 7,  # Sur NAS → On NAS ?
    7: 8,  # Parution terminée → Parution terminée
    # 8: Éditeur — not in ImportExcelService format, dropped
}

# Columns where Bibliotheque.xlsx takes priority on conflict
BIBLIO_PRIORITY_COLS = {2, 3, 4}  # Buy?, Last bought, Current
# Columns where nas-import.xlsx takes priority on conflict
NAS_PRIORITY_COLS = {6, 7}  # Last dled, On NAS ?
# Columns where we keep whichever is non-null (or Biblio if both set): 5, 8

SHEETS = ["BD", "Comics", "Mangas"]

# Sheet name → Type column value
SHEET_TYPE_VALUE = {
    "BD": "BD",
    "Comics": "Comics",
    "Mangas": "Manga",
    "Livre": "Livre",
}

# Tome extraction patterns (same as ImportBooksService)
TOME_PATTERNS = [
    r"^(.+?)\s*[-–]\s*[Tt]ome\s+(\d+)",
    r"^(.+?),\s*[Tt]ome\s+(\d+)",
    r"^(.+?)\s*[-–]\s*[Tt](\d+)\s",
    r"^(.+?)\s*[-–]\s*[Tt](\d+)$",
    r"^(.+?)\s+[Tt](\d+)\s*[-–:]",
    r"^(.+?)\s+[Tt](\d+)$",
    r"^(.+?),\s*[Tt](\d+)\s*[-–:]",
    r"^(.+?)\s*[-–]\s*n[oº°]?\s*(\d+)",
]

# Category → type value mapping
CATEGORY_TYPE_MAP = {
    "bd": "BD",
    "comics": "Comics",
    "manga": "Manga",
    "livre": "Livre",
}


def normalize(s):
    """Normalize title for exact key matching (same logic as ComicSeriesRepository)."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[-'\u2019.,!?():]+", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_deep(s):
    """Aggressive normalization for fuzzy dedup.

    Goes beyond normalize() to catch near-duplicates:
    - & and + → et, vs/versus unified
    - Remove $, [], {}
    - Strip ALL parenthetical/bracket content: (integrale), (T01-05), [Dark Horse]
    - Strip trailing version markers: V1, v01
    - Strip trailing date ranges: 2015-2016, 2017
    - Strip trailing issue ranges: 001-003, 01-06
    - Strip leading articles: l', le, la, les
    - Collapse spaces in single-letter sequences (acronyms)
    """
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    # & and + → et
    s = re.sub(r"[&+]", "et", s)
    # versus → vs
    s = s.replace("versus", "vs")
    # Remove $, [], {}
    s = re.sub(r"[$\[\]{}]", "", s)
    # Strip ALL parenthetical/bracket content BEFORE removing punctuation
    while re.search(r"\([^)]*\)", s):
        s = re.sub(r"\([^)]*\)", "", s)
    while re.search(r"\[[^\]]*\]", s):
        s = re.sub(r"\[[^\]]*\]", "", s)
    # Remove punctuation
    s = re.sub(r"[-'\u2019.,!?():]+", " ", s)
    # Strip leading articles
    s = re.sub(r"^(l |le |la |les )", "", s)
    # Strip trailing version/date/issue markers
    s = re.sub(r"\s+v\d+\s*$", "", s)
    s = re.sub(r"\s+\d{4}(-\d{4})?\s*$", "", s)
    s = re.sub(r"\s+\d{2,3}(-\d{2,3})?\s*$", "", s)
    # Strip noise words at end
    s = re.sub(r"\s+(integrale|omnibus|integrales?)\s*$", "", s)
    # Collapse single-letter sequences (acronyms): "W I L D C A T S" → "wildcats"
    s = re.sub(r"\b(\w) (\w) (\w)", lambda m: m.group(0).replace(" ", ""), s)
    # Collapse spaces
    s = re.sub(r"\s+", " ", s).strip()
    return s


def are_fuzzy_duplicates(key_a, key_b, title_a, title_b):
    """Check if two titles are near-duplicates using multiple strategies."""
    # Strategy 1: deep-normalized exact match
    deep_a = normalize_deep(title_a)
    deep_b = normalize_deep(title_b)
    if deep_a == deep_b:
        return True

    # Strategy 2: one is a prefix of the other (base series + variant)
    if len(deep_a) >= 5 and len(deep_b) >= 5:
        shorter, longer = sorted([deep_a, deep_b], key=len)
        if longer.startswith(shorter):
            extra = longer[len(shorter):].strip()
            extra_words = len(extra.split()) if extra else 0
            # Reject if extra is a roman numeral or arabic number (sequel)
            if re.match(r"^[ivxlc]+$|^\d+$", extra):
                pass
            # Reject if shorter is a single common word and extra adds a real word
            elif len(shorter.split()) == 1 and extra_words >= 1:
                pass
            # Accept if shorter is 60%+ of longer, or at most 1 noise extra word
            elif len(shorter) >= 12 and (len(shorter) >= len(longer) * 0.6 or extra_words <= 1):
                return True
            # Shorter multi-word titles: require 70%+ overlap
            elif len(shorter) >= 8 and len(shorter.split()) >= 2 and len(shorter) >= len(longer) * 0.7:
                return True

    # Strategy 3: high string similarity (catches typos, singular/plural)
    ratio = difflib.SequenceMatcher(None, deep_a, deep_b).ratio()
    # Higher threshold for short titles to avoid "Aster"→"Asterix", "Crusaders"→"Crusades"
    min_ratio = 0.92 if min(len(deep_a), len(deep_b)) < 12 else 0.85
    if ratio >= min_ratio and min(len(deep_a), len(deep_b)) >= 5:
        return True

    return False


def extract_series(title):
    """Extract series name and tome number from a book title."""
    for pat in TOME_PATTERNS:
        m = re.match(pat, title, re.UNICODE)
        if m:
            return m.group(1).strip(), int(m.group(2))
    return title.strip(), None


def detect_type_value(categories):
    """Determine the Type column value based on categories."""
    if not categories:
        return None
    cats = str(categories).lower()
    # Priority order: BD > Comics > Manga > Livre
    for keyword, type_val in [("bd", "BD"), ("comics", "Comics"), ("manga", "Manga"), ("livre", "Livre")]:
        if keyword in cats:
            return type_val
    return None


def clean_title(title):
    """Remove trailing dots and spaces from a title."""
    return title.rstrip(". ") if title else title


def read_sheet_rows(ws, max_cols=14):
    """Read all data rows from a worksheet, returning list of 15-element lists with type at index 0.

    The source sheets have 14 columns (no Type). This function reads them and returns
    15-element rows with None at index 0 (Type to be filled by caller).
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[0] if row else None
        if title is None:
            continue
        # Read source columns (14 cols), prepend None for Type
        values = list(row[:max_cols])
        while len(values) < max_cols:
            values.append(None)
        # Ensure title is string, clean trailing dots/spaces
        values[0] = clean_title(str(values[0]).strip()) if values[0] is not None else None
        # Prepend Type placeholder (index 0)
        values = [None] + values
        rows.append(values)
    return rows


def merge_rows(biblio_row, nas_row):
    """Merge two rows, taking the best value from each source."""
    merged = [None] * 15
    # Type: preserve from biblio_row
    merged[0] = biblio_row[0]
    # Title: prefer Bibliotheque (original user data)
    merged[1] = biblio_row[1]

    for col in range(2, 15):
        bval = biblio_row[col] if col < len(biblio_row) else None
        nval = nas_row[col] if col < len(nas_row) else None

        if bval is not None and nval is not None:
            # Both have values — apply priority rules
            if col in BIBLIO_PRIORITY_COLS:
                merged[col] = bval
            elif col in NAS_PRIORITY_COLS:
                merged[col] = nval
            else:
                merged[col] = bval  # default: Biblio wins
        else:
            # Take whichever is non-null
            merged[col] = bval if bval is not None else nval

    return merged


def nas_row_to_output(nas_row):
    """Convert a nas-import row to the output format (15 elements with Type at index 0)."""
    output = [None] * 15
    for nas_idx, out_idx in NAS_COL_MAP.items():
        if nas_idx < len(nas_row):
            output[out_idx] = nas_row[nas_idx]
    # Ensure title is string, clean trailing dots/spaces
    if output[1] is not None:
        output[1] = clean_title(str(output[1]).strip())
    return output


def parse_bought_value(val):
    """Parse 'Last bought' cell. Returns (specific_tomes: set|None, max_value: int|None, is_complete: bool)."""
    if val is None:
        return None, None, False
    s = str(val).strip().lower()
    if not s or s == "non":
        return None, None, False

    # "fini" or "fini N"
    m = re.match(r"fini\s*(\d+)?", s)
    if m:
        v = int(m.group(1)) if m.group(1) else None
        return None, v, True

    # "N+MHS"
    m = re.match(r"(\d+)\s*\+", s)
    if m:
        return None, int(m.group(1)), False

    # Comma-separated → specific tomes
    if "," in s:
        nums = {int(x.strip()) for x in s.split(",") if x.strip().isdigit() and int(x.strip()) > 0}
        return (nums if nums else None), (max(nums) if nums else None), False

    # Simple integer
    try:
        return None, int(float(s)), False
    except (ValueError, TypeError):
        return None, None, False


def format_specific_tomes(tomes):
    """Format a set of tome numbers as a CSV string."""
    return ", ".join(str(t) for t in sorted(tomes))


def format_tome_isbns(tome_isbns):
    """Format tome→ISBN mapping as 'ISBN1:T1,ISBN2:T8,...' string."""
    if not tome_isbns:
        return None
    return ",".join(f"{isbn}:T{num}" for num, isbn in sorted(tome_isbns.items()))


def main():
    # Verify all input files exist
    for path, name in [(BIBLIO_PATH, "Bibliotheque.xlsx"), (NAS_PATH, "nas-import.xlsx"), (LIVRES_PATH, "Livres.xlsx")]:
        if not os.path.exists(path):
            print(f"ERREUR : {name} introuvable → {path}")
            sys.exit(1)

    print("Chargement des fichiers sources...")
    wb_biblio = openpyxl.load_workbook(BIBLIO_PATH, read_only=True, data_only=True)
    wb_nas = openpyxl.load_workbook(NAS_PATH, read_only=True, data_only=True)
    wb_livres = openpyxl.load_workbook(LIVRES_PATH, read_only=True, data_only=True)

    # === Step 1: Extract bought tomes from Livres.xlsx ===
    print("\n--- Extraction des tomes achetés depuis Livres.xlsx ---")
    ws_livres = wb_livres[wb_livres.sheetnames[0]]

    # series_key → {type_value, max_tome, name, count, tome_isbns, author, cover, publisher, categories, description}
    bought_tomes = {}
    for row in ws_livres.iter_rows(min_row=2, values_only=True):
        title = row[1]
        categories = row[5]
        if not title:
            continue
        series_name, tome_num = extract_series(str(title))
        type_value = detect_type_value(categories)
        key = normalize(series_name)

        # Clean ISBN: remove .0 suffix from Excel
        isbn = row[0]
        if isbn is not None:
            isbn_str = str(isbn)
            if isbn_str.endswith(".0"):
                isbn_str = isbn_str[:-2]
            isbn = isbn_str if isbn_str else None

        # Clean cover: replace file:// URLs with None
        cover = row[4]
        if cover and str(cover).startswith("file://"):
            cover = None

        if key not in bought_tomes:
            tome_isbns = {}
            if isbn and tome_num is not None:
                tome_isbns[tome_num] = isbn
            bought_tomes[key] = {
                "name": series_name,
                "type_value": type_value,
                "max_tome": tome_num,
                "specific_tomes": {tome_num} if tome_num else set(),
                "count": 1,
                "tome_isbns": tome_isbns,
                "author": row[2],
                "cover": cover,
                "publisher": row[3],
                "categories": categories,
                "description": row[6],
            }
        else:
            bought_tomes[key]["count"] += 1
            if tome_num is not None:
                bought_tomes[key]["specific_tomes"].add(tome_num)
                current_max = bought_tomes[key]["max_tome"]
                if current_max is None or tome_num > current_max:
                    bought_tomes[key]["max_tome"] = tome_num
                if isbn:
                    bought_tomes[key]["tome_isbns"][tome_num] = isbn
            if bought_tomes[key]["author"] is None and row[2]:
                bought_tomes[key]["author"] = row[2]
            if bought_tomes[key]["cover"] is None and cover:
                bought_tomes[key]["cover"] = cover
            if bought_tomes[key]["publisher"] is None and row[3]:
                bought_tomes[key]["publisher"] = row[3]
            if bought_tomes[key]["categories"] is None and categories:
                bought_tomes[key]["categories"] = categories
            if bought_tomes[key]["description"] is None and row[6]:
                bought_tomes[key]["description"] = row[6]

    print(f"  {len(bought_tomes)} séries extraites ({sum(1 for v in bought_tomes.values() if v['max_tome'])} avec tomes)")

    # === Step 2: Merge Bibliotheque + NAS per sheet, collect all rows ===
    all_rows = []  # All rows with Type at index 0
    stats = {}
    enrichment_stats = {"matched": 0, "updated_bought": 0}

    for sheet_name in SHEETS:
        print(f"\n--- Fusion onglet {sheet_name} ---")
        type_value = SHEET_TYPE_VALUE[sheet_name]

        # Read both sources
        biblio_rows = read_sheet_rows(wb_biblio[sheet_name]) if sheet_name in wb_biblio.sheetnames else []
        nas_rows = []
        if sheet_name in wb_nas.sheetnames:
            ws_nas = wb_nas[sheet_name]
            for row in ws_nas.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    converted = nas_row_to_output(list(row))
                    nas_rows.append(converted)

        # Set type for all source rows
        for row in biblio_rows:
            row[0] = type_value
        for row in nas_rows:
            row[0] = type_value

        # Index by normalized title
        biblio_index = {}
        for row in biblio_rows:
            key = normalize(row[1])
            if key:
                biblio_index[key] = row

        nas_index = {}
        for row in nas_rows:
            key = normalize(row[1])
            if key:
                nas_index[key] = row

        # Merge
        merged_rows = []
        merged_keys = set()
        both_count = 0
        only_biblio = 0
        only_nas = 0

        # Process Bibliotheque rows first (preserving order)
        for row in biblio_rows:
            key = normalize(row[1])
            if not key or key in merged_keys:
                continue
            merged_keys.add(key)

            if key in nas_index:
                merged_rows.append(merge_rows(row, nas_index[key]))
                both_count += 1
            else:
                padded = list(row[:15])
                while len(padded) < 15:
                    padded.append(None)
                merged_rows.append(padded)
                only_biblio += 1

        # Add NAS-only rows
        for row in nas_rows:
            key = normalize(row[1])
            if not key or key in merged_keys:
                continue
            merged_keys.add(key)
            merged_rows.append(row)
            only_nas += 1

        # Enrich with Livres.xlsx bought tomes
        enriched_in_sheet = 0
        merged_by_key = {normalize(r[1]): r for r in merged_rows}

        def find_merged_row(livres_key):
            """Find matching row: exact match first, then check if any merged key starts with livres_key or vice versa."""
            if livres_key in merged_by_key:
                return merged_by_key[livres_key]
            # Fuzzy: livres_key is a prefix of a merged key, or merged key is a prefix of livres_key
            for mkey, mrow in merged_by_key.items():
                if len(livres_key) >= 4 and len(mkey) >= 4:
                    if mkey.startswith(livres_key) or livres_key.startswith(mkey):
                        return mrow
            return None

        # Match Livres.xlsx entries to this sheet's type
        # For "Mangas" sheet, match type_value "Manga" from Livres.xlsx
        livres_match_types = {type_value}
        if sheet_name == "Mangas":
            livres_match_types.add("Manga")

        for key, info in bought_tomes.items():
            if info["type_value"] not in livres_match_types:
                continue
            row = find_merged_row(key)
            if row is None:
                continue
            enrichment_stats["matched"] += 1
            info["_matched"] = True

            # Update "Last bought" (col 3) with specific tomes from Livres.xlsx
            if info["max_tome"] is not None:
                existing_specific, existing_max, existing_complete = parse_bought_value(row[3])

                if existing_complete:
                    # Already "fini" — all tomes bought, nothing to add
                    pass
                elif existing_specific is not None:
                    # Existing is a CSV list — merge the specific tomes
                    merged_tomes = existing_specific | info["specific_tomes"]
                    row[3] = format_specific_tomes(merged_tomes)
                    if merged_tomes != existing_specific:
                        enrichment_stats["updated_bought"] += 1
                        enriched_in_sheet += 1
                elif existing_max is not None:
                    # Existing is a single number (= all tomes 1..N bought)
                    # Add Livres.xlsx tomes that are beyond existing_max
                    extra = {t for t in info["specific_tomes"] if t > existing_max}
                    if extra:
                        # Convert to specific list: 1..existing_max + extras
                        all_tomes = set(range(1, existing_max + 1)) | extra
                        row[3] = format_specific_tomes(all_tomes)
                        enrichment_stats["updated_bought"] += 1
                        enriched_in_sheet += 1
                else:
                    # No existing value — set specific tomes from Livres.xlsx
                    row[3] = format_specific_tomes(info["specific_tomes"])
                    enrichment_stats["updated_bought"] += 1
                    enriched_in_sheet += 1

            # Set Buy? = "oui" if not already set
            if row[2] is None:
                row[2] = "oui"

            # Populate metadata cols 9-14 from Livres.xlsx
            if row[9] is None and format_tome_isbns(info["tome_isbns"]):
                row[9] = format_tome_isbns(info["tome_isbns"])
            if row[10] is None and info["author"]:
                row[10] = info["author"]
            if row[11] is None and info["cover"]:
                row[11] = info["cover"]
            if row[12] is None and info["publisher"]:
                row[12] = info["publisher"]
            if row[13] is None and info["categories"]:
                row[13] = info["categories"]
            if row[14] is None and info["description"]:
                row[14] = info["description"]

        all_rows.extend(merged_rows)

        total = len(merged_rows)
        stats[sheet_name] = {
            "total": total,
            "both": both_count,
            "only_biblio": only_biblio,
            "only_nas": only_nas,
            "enriched": enriched_in_sheet,
        }
        print(f"  {total} séries (commun: {both_count}, biblio seul: {only_biblio}, NAS seul: {only_nas}, enrichis Livres: {enriched_in_sheet})")

    # === Step 3: Process Livre sheet from Bibliotheque ===
    print("\n--- Onglet Livre ---")
    livre_rows = read_sheet_rows(wb_biblio["Livre"]) if "Livre" in wb_biblio.sheetnames else []
    for row in livre_rows:
        row[0] = "Livre"

    # Enrich Livre rows with Livres.xlsx
    livre_by_key = {normalize(r[1]): r for r in livre_rows}
    enriched_livre = 0
    for key, info in bought_tomes.items():
        if info["type_value"] != "Livre":
            continue
        if key not in livre_by_key:
            continue
        row = livre_by_key[key]
        enrichment_stats["matched"] += 1
        info["_matched"] = True
        if info["max_tome"] is not None:
            existing_specific, existing_max, existing_complete = parse_bought_value(row[3])
            if not existing_complete:
                if existing_specific is not None:
                    merged_tomes = existing_specific | info["specific_tomes"]
                    if merged_tomes != existing_specific:
                        row[3] = format_specific_tomes(merged_tomes)
                        enrichment_stats["updated_bought"] += 1
                        enriched_livre += 1
                elif existing_max is not None:
                    extra = {t for t in info["specific_tomes"] if t > existing_max}
                    if extra:
                        all_tomes = set(range(1, existing_max + 1)) | extra
                        row[3] = format_specific_tomes(all_tomes)
                        enrichment_stats["updated_bought"] += 1
                        enriched_livre += 1
                else:
                    row[3] = format_specific_tomes(info["specific_tomes"])
                    enrichment_stats["updated_bought"] += 1
                    enriched_livre += 1
        if row[2] is None:
            row[2] = "oui"

        # Populate metadata cols 9-14 from Livres.xlsx
        if row[9] is None and format_tome_isbns(info["tome_isbns"]):
            row[9] = format_tome_isbns(info["tome_isbns"])
        if row[10] is None and info["author"]:
            row[10] = info["author"]
        if row[11] is None and info["cover"]:
            row[11] = info["cover"]
        if row[12] is None and info["publisher"]:
            row[12] = info["publisher"]
        if row[13] is None and info["categories"]:
            row[13] = info["categories"]
        if row[14] is None and info["description"]:
            row[14] = info["description"]

    # Add unmatched Livres.xlsx entries to Livre rows
    unmatched_added = 0
    for key, info in bought_tomes.items():
        if info.get("_matched"):
            continue
        # New row: type + title + metadata in cols 9-14, tracking cols 2-8 empty
        new_row = [None] * 15
        new_row[0] = "Livre"
        new_row[1] = info["name"]
        new_row[9] = format_tome_isbns(info["tome_isbns"])
        new_row[10] = info["author"]
        new_row[11] = info["cover"]
        new_row[12] = info["publisher"]
        new_row[13] = info["categories"]
        new_row[14] = info["description"]
        livre_rows.append(new_row)
        unmatched_added += 1

    all_rows.extend(livre_rows)

    stats["Livre"] = {"total": len(livre_rows), "enriched": enriched_livre, "unmatched_added": unmatched_added}
    print(f"  {len(livre_rows)} séries (enrichis Livres: {enriched_livre}, ajoutés non-matchés: {unmatched_added})")

    # === Step 3b: Deduplicate cross-type entries ===
    print("\n--- Déduplication cross-types ---")
    type_priority = {"BD": 1, "Manga": 2, "Comics": 3, "Livre": 4}

    # Build index of all titles
    all_entries = {}  # normalized_key -> list of (index, type, has_data)
    for idx, row in enumerate(all_rows):
        title = str(row[1]).strip() if row[1] else ""
        key = normalize(title)
        if not key:
            continue
        has_data = any(v is not None for v in row[2:15])
        all_entries.setdefault(key, []).append({
            "idx": idx,
            "type": row[0],
            "has_data": has_data,
            "priority": type_priority.get(row[0], 99),
        })

    indices_to_delete = set()
    cross_dupes = 0

    for key, entries in all_entries.items():
        types = set(e["type"] for e in entries)
        if len(types) <= 1:
            continue

        # Keep entry with data + best priority; merge data into keeper
        entries.sort(key=lambda e: (not e["has_data"], e["priority"]))
        keeper = entries[0]

        for other in entries[1:]:
            # Merge non-null values from other into keeper
            for col in range(2, 15):
                if all_rows[keeper["idx"]][col] is None:
                    other_val = all_rows[other["idx"]][col]
                    if other_val is not None:
                        all_rows[keeper["idx"]][col] = other_val

            indices_to_delete.add(other["idx"])
            cross_dupes += 1
            title = all_rows[other["idx"]][1]
            print(f"  '{title}' [{other['type']}] fusionné dans [{keeper['type']}]")

    # Remove duplicates (reverse order to preserve indices)
    for idx in sorted(indices_to_delete, reverse=True):
        del all_rows[idx]

    if cross_dupes:
        print(f"  {cross_dupes} doublons cross-types fusionnés")
    else:
        print("  Aucun doublon cross-types")

    # Sort all rows by type priority then title
    all_rows.sort(key=lambda r: (type_priority.get(r[0], 99), normalize(r[1])))

    # === Write single "Import" sheet ===
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Import"
    ws_out.append(BIBLIO_HEADERS)
    for row in all_rows:
        ws_out.append(row)

    wb_out.save(OUTPUT_MERGED)
    print(f"\n✓ Fichier fusionné : {OUTPUT_MERGED}")

    # Close source workbooks
    wb_biblio.close()
    wb_nas.close()
    wb_livres.close()

    # === Summary ===
    print("\n=== RÉSUMÉ ===")
    print(f"Total séries dans import.xlsx : {len(all_rows)}")
    for sheet_name, s in stats.items():
        print(f"  {sheet_name}: {s['total']}")
    print(f"Enrichissements Livres.xlsx : {enrichment_stats['matched']} matchés, {enrichment_stats['updated_bought']} 'Last bought' mis à jour")


if __name__ == "__main__":
    main()
