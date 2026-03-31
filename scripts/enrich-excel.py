#!/usr/bin/env python3
"""
Enrichit import.xlsx produit par merge-excel.py :
1. Parution "fini" sans nombre → cherche le nombre total de tomes sur Wikipedia
2. Séries achetées sans ISBN → cherche l'ISBN du 1er tome sur Google Books (col 9)

Format : feuille unique "Import" avec colonne Type en col 0.

Filtre strict : un résultat n'est accepté que si le titre trouvé
partage suffisamment de mots significatifs avec le titre recherché.
"""

import json
import os
import re
import sys
import time
import unicodedata
import urllib.parse
import urllib.request

import openpyxl

# --- Paths ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MERGED_PATH = os.path.join(BASE_DIR, "var", "import.xlsx")

HEADERS = {"User-Agent": "BibliothequeBot/1.0 (bibliotheque merge-excel enrichment)"}

TYPE_KEYWORDS = {
    "BD": "bande dessinée",
    "Comics": "comics",
    "Manga": "manga",
    "Livre": "roman",
}

# Mots trop génériques pour la comparaison de titres
STOP_WORDS = {
    "le", "la", "les", "l", "de", "du", "des", "d", "un", "une",
    "et", "en", "au", "aux", "a", "the", "of", "and", "in", "on",
    "tome", "volume", "n", "no", "t", "bd", "manga", "comics",
    "serie", "integrale", "edition", "nouvelle",
}


def normalize(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[-'\u2019.,!?():{}\"]+", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def significant_words(title):
    """Extract significant words (3+ chars, not stop words) from a title."""
    words = set()
    for w in normalize(title).split():
        if len(w) >= 3 and w not in STOP_WORDS:
            words.add(w)
    return words


def titles_match(searched_title, found_title):
    """Check if found_title is a plausible match for searched_title.

    For 1-word titles: the word must appear in the found title.
    For 2-word titles: both words must appear.
    For 3+ word titles: at least 50% of significant words must match.
    """
    searched_words = significant_words(searched_title)
    found_words = significant_words(found_title)

    if not searched_words:
        return False

    overlap = searched_words & found_words
    word_count = len(searched_words)

    if word_count <= 2:
        # Short titles: all significant words must match
        return len(overlap) == word_count

    # Longer titles: at least 50%
    return len(overlap) >= (word_count + 1) // 2


# ===================================================================
# Wikipedia: volume count lookup
# ===================================================================

def wiki_search(query, limit=3):
    """Search French Wikipedia, return list of result titles."""
    url = (
        "https://fr.wikipedia.org/w/api.php?action=query&list=search"
        f"&srsearch={urllib.parse.quote(query)}&format=json&srlimit={limit}"
    )
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read())
        return [r["title"] for r in data.get("query", {}).get("search", [])]
    except Exception:
        return []


def wiki_get_wikitext(title):
    """Get wikitext content of a Wikipedia article."""
    url = (
        "https://fr.wikipedia.org/w/api.php?action=parse"
        f"&page={urllib.parse.quote(title)}&prop=wikitext&format=json"
    )
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read())
        return data.get("parse", {}).get("wikitext", {}).get("*", "")
    except Exception:
        return ""


def extract_volume_count(wikitext):
    """Extract volume/album count from Wikipedia infobox wikitext."""
    patterns = [
        r"[Nn]ombre[_ ]d['\u2019]albums?\s*=\s*(\d+)",
        r"[Nn]b[_ ]albums?\s*=\s*(\d+)",
        r"[Nn]ombre[_ ]de[_ ]volumes?\s*=\s*(\d+)",
        r"[Nn]b[_ ]volumes?\s*=\s*(\d+)",
        r"[Vv]olumes?\s*=\s*(\d+)",
        r"[Nn]ombre[_ ]tomes?\s*=\s*(\d+)",
    ]
    for pat in patterns:
        m = re.search(pat, wikitext)
        if m:
            return int(m.group(1))
    return None


def lookup_volume_count(series_name, type_keyword):
    """Search Wikipedia for a series and return its volume count.

    Tries multiple search queries and checks title similarity to avoid
    false positives (e.g. "Incantations" matching "La Quête de l'oiseau du temps").
    """
    queries = [
        f"{series_name} {type_keyword}",
        series_name,
    ]

    for query in queries:
        results = wiki_search(query, limit=3)
        for wiki_title in results:
            # Filter: article title must share significant words with series name
            if not titles_match(series_name, wiki_title):
                continue

            wikitext = wiki_get_wikitext(wiki_title)
            count = extract_volume_count(wikitext)
            if count is not None and count <= 200:  # Sanity: no series has 200+ tomes
                return count, wiki_title
            time.sleep(0.3)

    return None, None


# ===================================================================
# Google Books: ISBN lookup
# ===================================================================

def gbooks_search(query, max_results=5):
    """Search Google Books and return list of (isbn, title, authors) tuples."""
    url = (
        "https://www.googleapis.com/books/v1/volumes"
        f"?q={urllib.parse.quote(query)}&maxResults={max_results}&langRestrict=fr"
    )
    req = urllib.request.Request(url, headers=HEADERS)
    results = []
    try:
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read())
        for item in data.get("items", []):
            vi = item["volumeInfo"]
            isbn = None
            for ident in vi.get("industryIdentifiers", []):
                if ident["type"] == "ISBN_13":
                    isbn = ident["identifier"]
                    break
            if isbn:
                results.append((isbn, vi.get("title", ""), vi.get("authors", [])))
    except Exception:
        pass
    return results


def lookup_isbn(series_name, tome_number=1):
    """Try multiple strategies to find ISBN for a series tome.

    Validates that the Google Books result title matches the series name
    to avoid false positives (e.g. "Segments" matching an anatomy book).
    """
    strategies = [
        f'intitle:"{series_name}" intitle:"tome {tome_number}"',
        f'intitle:"{series_name}" intitle:"t{tome_number}"',
        f'intitle:"{series_name}" intitle:"n°{tome_number}"',
        f'"{series_name}" "tome {tome_number}"',
        f'intitle:"{series_name}"',
    ]

    for query in strategies:
        results = gbooks_search(query, max_results=5)
        for isbn, book_title, authors in results:
            # Filter: book title must share significant words with series name
            if titles_match(series_name, book_title):
                return isbn, book_title
        time.sleep(0.5)

    return None, None


# ===================================================================
# Main
# ===================================================================

def main():
    if not os.path.exists(MERGED_PATH):
        print(f"ERREUR : import.xlsx introuvable → {MERGED_PATH}")
        print("Lancez d'abord: python3 scripts/merge-excel.py")
        sys.exit(1)

    # === Task 1: Enrich "fini" series with volume count from Wikipedia ===
    print("=== Enrichissement des parutions « fini » via Wikipedia ===\n")

    wb_merged = openpyxl.load_workbook(MERGED_PATH)
    wiki_found = 0
    wiki_not_found = []
    wiki_total = 0

    ws = wb_merged["Import"]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        type_value = str(row[0].value).strip() if row[0].value else ""
        type_keyword = TYPE_KEYWORDS.get(type_value, "")

        parution_cell = row[5]  # Column F = Parution (0-indexed col 5)
        parution = str(parution_cell.value).strip().lower() if parution_cell.value is not None else ""

        if "fini" not in parution:
            continue

        # Already has a number?
        if re.search(r"\d+", parution):
            continue

        title = str(row[1].value).strip() if row[1].value else ""
        if not title:
            continue

        wiki_total += 1
        count, wiki_title = lookup_volume_count(title, type_keyword)
        time.sleep(0.3)  # Rate limit

        if count is not None:
            ws.cell(row=row_idx, column=6).value = f"fini {count}"  # Column F (1-indexed = 6)
            wiki_found += 1
            print(f"  ✓ [{type_value}] {title} → {count} tomes (via {wiki_title})")
        else:
            wiki_not_found.append((type_value, title))
            print(f"  ✗ [{type_value}] {title} → non trouvé")

    wb_merged.save(MERGED_PATH)
    print(f"\nWikipedia : {wiki_found}/{wiki_total} trouvés")
    if wiki_not_found:
        print(f"Non trouvés ({len(wiki_not_found)}) :")
        for type_val, title in wiki_not_found:
            print(f"  [{type_val}] {title}")

    # === Task 2: Add ISBNs for bought series without ISBN in import.xlsx ===
    print("\n=== Recherche d'ISBN via Google Books ===\n")

    # Reload merged file (freshly saved)
    wb_merged = openpyxl.load_workbook(MERGED_PATH)

    # Find bought series without ISBN (col 9, 0-indexed)
    bought_series = []
    ws = wb_merged["Import"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        type_value = str(row[0].value).strip() if row[0].value else ""
        title = str(row[1].value).strip() if row[1].value else ""
        buy_status = str(row[2].value).strip().lower() if row[2].value is not None else ""
        bought = str(row[3].value).strip() if row[3].value is not None else ""
        isbn_cell = row[9] if len(row) > 9 else None  # Column J = ISBN (0-indexed col 9)
        existing_isbn = isbn_cell.value if isbn_cell else None

        has_bought = bought != "" and bought.lower() != "non"
        has_buy = buy_status in ("oui", "fini")

        if (has_bought or has_buy) and not existing_isbn:
            # Determine first bought tome number
            first_tome = 1
            if "," in bought:
                nums = [int(x.strip()) for x in bought.split(",") if x.strip().isdigit() and int(x.strip()) > 0]
                if nums:
                    first_tome = min(nums)

            bought_series.append((type_value, row_idx, title, first_tome))

    isbn_found = 0
    isbn_not_found = []

    for type_value, row_idx, title, first_tome in bought_series:
        isbn, book_title = lookup_isbn(title, first_tome)
        time.sleep(0.3)  # Rate limit

        if isbn:
            ws = wb_merged["Import"]
            ws.cell(row=row_idx, column=10).value = isbn  # Column J = ISBN (1-indexed = 10)
            isbn_found += 1
            print(f"  ✓ [{type_value}] {title} T{first_tome} → {isbn} ({book_title})")
        else:
            isbn_not_found.append((type_value, title, first_tome))
            print(f"  ✗ [{type_value}] {title} T{first_tome} → non trouvé")

    wb_merged.save(MERGED_PATH)
    print(f"\nGoogle Books : {isbn_found}/{len(bought_series)} ISBN trouvés")
    if isbn_not_found:
        print(f"Non trouvés ({len(isbn_not_found)}) :")
        for type_val, title, tome in isbn_not_found:
            print(f"  [{type_val}] {title} T{tome}")

    # === Summary ===
    print("\n=== RÉSUMÉ ===")
    print(f"Wikipedia (parutions fini) : {wiki_found}/{wiki_total} enrichis")
    print(f"Google Books (ISBN)        : {isbn_found}/{len(bought_series)} ajoutés à import.xlsx")


if __name__ == "__main__":
    main()
